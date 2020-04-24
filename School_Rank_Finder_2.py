# -*- coding: utf-8 -*-
"""
Created on Tue Mar 31 07:26:16 2020

@author: decla
"""


import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import openpyxl
import xlwt
import numpy
import math
from sklearn.linear_model import LinearRegression
from sympy import symbols, solve, diff
import pandas
from datetime import datetime 

#some global data objects
assessments= {}
curricula = {}
#schools, classes, learners, results, etc..

# Load the Excel file and access the Maths and Language results sheets.
#file = input('Enter the file directory')
file = input("File Name")
workbook = openpyxl.load_workbook(file, data_only=True)

# Read in the structured data in this format:
'''
    data = {Grade 1:{ <--put Grade at the root
            {School 1:{
                clazz 1: 
                    Students:{
                        Student 1:{ 
                            First Name:___, 
                            Surname:___, 
                            S/No:___,
                            Oldest:___, 
                            Most Recent:___, 
                            Results: {
                                Assessments: {
                                    Assessment: {
                                        Overall: mark,
                                        Attempted: mark,
                                        Curriculum: {
                                            
                                            Topic 1: { <--Content Domain
                                                subtopic 1: mark
                                                subtopic 2: mark
                                                ...
                                                subtopic 'n': mark
                                                }
                                            Topic 2: { <---Cognitive Domain
                                                subtopic 1: mark
                                                subtopic 2: mark
                                                ...
                                                subtopic 'n': mark
                                                }
                                            Topic 3: { <-- Grade level
                                                subtopic 1: mark
                                                subtopic 2: mark
                                                ...
                                                subtopic 'n': mark
                                                }
                                        }
                                    }
                                }
                            }
                            
                            Scores:{
                                    Q1:{
                                        Mark:___,
                                        Grade Level:___,
                                        Cognitive Domain:___,
                                        Content Domain:___}
                                    Q2:{
                                    .
                                    .
                                    .
                                    }}}
                        Student 2:{
                            .
                            .
                            .}}}
                clazz 2:{
                    .
                    .
                    .
                    }}
            School 2:{
                .
                .
                .
                }}
            Grade 2:{
                .
                .
                .}}

'''
# Input the raw data from a given sheet
def input_data(assessment_name, headings_row, rankthreshold):
    
    sheet = workbook[assessment_name]
    
    #TBD: curricula by subject, not by assessment...
    assessments[assessment_name]={}
    curricula[assessment_name]={}
    
    # Input the raw data
    
    # Fetch the headings and configure the curriculum and assessment data
    next_cell = ''
    headings = [] #TBD: Not necessarily unique
    # contentTopics = []
    # cognitiveTopics = []
    questions = {}
    assessmentmark = 0.00
    count = 1
    nutopics = 0
    topiccol = 0

    while next_cell != None:
        next_cell = sheet.cell(row=headings_row,column=count).value
        headings.append(next_cell)
        
        #if it's a question column update the questions list and topic maps
        if next_cell[0] == 'Q' and next_cell[1].isnumeric():
            #how many topics in curriculum?
            # if next_cell[1]=='1' :
            #while we are still on question 1
            if topiccol == 0 :
                #once we have found our Questions: set the Curriculum Topic col one back to get the Topic names
                topiccol=count-1
                nexttopic = 1
                #get the first Topic , if it exists.
                cell_above = sheet.cell(row=headings_row - 1,column=count).value
                while cell_above !=next_cell:
                    #if we found a valid subtopc and it's not the Questin aagain, add the Topic to the Curricula
                    topicname = sheet.cell(row=headings_row - nexttopic,column=topiccol).value
                    if topicname != 'Weight/mark':
                        #if its not the mark, then add the topic to the curriculua with empty list of subtopics, 
                        #and store the total marks in the assessment topics {topicname : topicmarks}
                        curricula[assessment_name][topicname]={}
                        nutopics +=1
                            
                    nexttopic +=1
                    #get the next subtopic
                    cell_above = sheet.cell(row=headings_row - nexttopic,column=count).value
                     
            #get the question attributes
            qmark = sheet.cell(row=headings_row-(nutopics+1),column=count).value
            questions[next_cell]={}
            questions.update({next_cell:{'mark': qmark}})           
            questions[next_cell]['subtopics']=[]
            for t in range(1, nutopics+1):
                topicname = sheet.cell(row=headings_row-t,column=topiccol).value
                subtopicname = sheet.cell(row=headings_row-t,column=count).value
                
                #complete the curricula subtopc to the topic if not there already
                # if curricula[topicname] == None :
                #     curricula[topicname].append(subtopicname)
                    
                if subtopicname not in curricula[assessment_name][topicname]:
                    curricula[assessment_name][topicname][subtopicname]=qmark
                else:
                    curricula[assessment_name][topicname][subtopicname]+=qmark
                        
                #add to the total marks in this topic in this assessment
                # subtopictotalmark += mark
               

                #associate question to it's subtopics
                questions[next_cell]['subtopics'].append(subtopicname)
            
            assessmentmark += qmark
            
            

        #get next heading    
        count+=1
        next_cell = sheet.cell(row=headings_row,column=count).value
    
    # questions.update({'totalmark':assessmentmark})    
    assessments.update({assessment_name:{'questions':questions,'totalmarks':assessmentmark}})
    # questions.update({next_cell:{'mark': qmark}})
    # print (assessments)
    # print (curricula)
    
    # read in all the learners, their scores and calculate averages, rannks for learners, classes, school and grade
    data={}
    current_cell=''
    col_count=headings_row
    
    #start at the Q1 column and row = headings row+1 (where the student starts)
    row_count = headings_row+1
    student = sheet.cell(row = row_count , column = headings.index('S/No')+1).value
    
    while student != None:
        #get the learner details
        grade = sheet.cell(row = row_count , column = headings.index('Grade')+1).value
        school = sheet.cell(row = row_count , column = headings.index('School')+1).value
        clazz = sheet.cell(row = row_count , column = headings.index('Class')+1).value
        
        data.update(grade {school: {clazz: {student:{'Grade':grade, 'School':school, 'Class':clazz, 'Results':{}},'Results'={}},'Results'={}},'Results'={}})
        
        totalmarks = 0.00
        questionmark =0.00
        overall = 0.00
        attempted = 0.00
        studentmark = 0.00
        unattemptedmarks = 0.00
        
            
        for qname , q in questions.items() :
            
            col_count = headings.index(qname)+1
           #get the question mark
            qmark = q['mark']
            totalmarks +=qmark
            
            #get the stedent score for question
            cellvalue = sheet.cell(row = row_count, column = col_count).value
            
            #if the results is a non-attempt of "-" then increment the non attemptes, and mark it wrong
            if cellvalue == '-':
                unattemptedmarks += qmark
                mark = 0
            else:
                #set the question mark
                mark = cellvalue
                
                
            #increment this students totoal mark
            studentmark += mark
              
            
            #for the linked Question Topics: Take the student mark and update the data table marks for all topics, overalls etc..            
            for subtopicname in q['subtopics']:
                # data[grade][school][clazz][student]['Results'][subtopicname)]['totalmarks']+= qmark
                # data[grade][school][clazz]['Results'][subtopicname]['totalmarks']+=qmark
                # data[grade][school]['Results'][subtopicname]['totalmarks']+=qmark
                # data[grade]['Results'][subtopicname]['totalmarks']+=qmark
                
                data[grade][school][clazz][student]['Results'][subtopicname)]['studentmarks']+= mark
                data[grade][school][clazz]['Results'][subtopicname]['studentmarks']+=mark
                data[grade][school]['Results'][subtopicname]['studentmarks']+=mark
                data[grade]['Results'][subtopicname]['studentmarks']+=mark
                
                data[grade][school][clazz][student]['Results'][subtopicname)]['unattempted']+= mark
                data[grade][school][clazz]['Results'][subtopicname]['unattempted']+=mark
                data[grade][school]['Results'][subtopicname]['unattempted']+=mark
                data[grade]['Results'][subtopicname]['unattempted']+=mark
            
                                    
        #set the STUDENT topic %'s and if Grade Level, the rank
        #for the students collection of results
        for results in data[grade][school][clazz][student]['Results']:
            #sort the topics (in particular we want the Grades ordered in Ascencing order)
            # sortedkeys = sorted(list(results))
            ranktopiccount = 1
            #for each subresult by topic
            # for stopic in sorted(list(results))sortedkeys:
            for stopic, result in sorted(results):
                # result = results(topickey)
                
                #Get the Topic for the resultsubtopic
                for topickey, subtopicitem in curriculu[assessment_name]:
                    if subtopicitem == stopic:
                        currtopic = topickey
                        #found the topic; exit
                        break
                 
                #total marks for the subtopic in this assessment
                totalmarks = curricula[assessment_name][currtopic][stopic]                                
                # totalmarks = result['totalmarks']
                studentmarks = result['studentmarks']
                unattemptedmarks = result['unattemptedmarks']
                
                
                studentscore = studentmarks/totalmarks * 100
                attemptedscore = studentmarks/(totalmarks - unattemptedmarks) * 100
                
                #now save the overall and attepted for the subtopic
                results.append('overall': studentscore)
                results.append('attempted': attemptedscore)
                 
                #is the topic a Grade Level? Calculate the Grade Rank
                if currtopic = 'Grade level':
                    #store the first topic as the grade rank to start anyway, irrespective of the score. If we get anything higher, we'll replace it
                    if ranktopiccount == 1:
                        results['graderank']=stopic
                    else
                        #if the score is >= threshold, and the stopic is a higher grade then the current, the replace the rank
                        if studentscore >= threshold:
                        
                            #is the scope > threshold? Save this
                            currrank = results['graderank']
                            
                            #we shouldn'thave to check here that the next grade is higher (>) as it's sorted, right?
                            # if stopic > currrank:
                            
                            #overwrite the graderank to this Grade Level subtopic
                            results['graderank']=stopic
                            
                            #increment the class, school and Grade grade rank values (nu. learners in the rank)
                            data[grade][school][clazz]['Results'][subtopicname]['graderankcount']+=1
                            data[grade][school]['Results'][subtopicname]['graderankcount']+=1
                            data[grade]['Results'][subtopicname]['graderankcount']+=1
                        
                    #increment our grade rank topic count
                    ranktopiccount+=1

                        
        #set the STUDENT overall average, attempted_average and the curriculum topic averages  
        totalassessmentmarks = assessments[assessment_name]['totalmarks']
        data[school][grade][clazz][student]['Results']['Overall'] = studentmark/totalmarks * 100
        data[school][grade][clazz][student]['Results']['Attempted'] = studentmark/(totalassessmentmarks - unattemptedmarks) * 100
                       
        #TODO: Count Nu. students in grade/school/class
        data[grade][school][clazz]['nustudents']+=1
        data[grade][school]['nustudents']+=1
        data[grade]['nustudents']+=1
        
        #get the next student
        row_count += 1
        student = sheet.cell(row = row_count , column = headings.index('S/No')+1).value
        
    #now set class, school, grade results
    for grade in data:
        studenttotalmarks = grade['studentmarks']
        studentunattemptedmarks =grade['unattemptedmarks']
        gradenustudents = data[grade]['nustudents']
        gradeoverall = (studenttotalmarks/(totalassessmentmarks * gradenustudents) * 100)
        grade['Results']['Overall']=gradeoverall


    return data

# Rank each student according to grade/cognitive/content
def rank_students(data,rank,threshold):
    ranks=[]
    student_ranks = {}
    student_scores={}
    student_averages={}
    for school in data:
        student_scores[school]={}
        for grade in data[school]:
            for clazz in data[school][grade]:
                for student in data[school][grade][clazz]:
                    if student != 'Teacher':
                        student_scores[school][student]={'Details':data[school][grade][clazz][student]}
                        for question in data[school][grade][clazz][student]['Scores']:
                            if data[school][grade][clazz][student]['Scores'][question][rank] not in student_scores[school][student]:
                                student_scores[school][student].update({data[school][grade][clazz][student]['Scores'][question][rank]:[data[school][grade][clazz][student]['Scores'][question]['Mark']]})
                            else:
                                student_scores[school][student][data[school][grade][clazz][student]['Scores'][question][rank]].append(int(data[school][grade][clazz][student]['Scores'][question]['Mark']))
    for school in student_scores:
        student_averages[school]={}
        for student in student_scores[school]:
            if student != 'Teacher':
                student_averages[school][student]={'Details':student_scores[school][student]['Details']}
                for rank in student_scores[school][student]:
                    if rank!='Details':
                        if rank not in ranks:
                            ranks.append(rank)
                            # STEVE: Learner averages are not matching the XL
                        student_averages[school][student][rank]=sum(student_scores[school][student][rank])/len(student_scores[school][student][rank])*100                    
    for school in student_averages:
        student_ranks[school]={}
        for student in student_averages[school]:
            student_ranks[school][student]={'Details':student_averages[school][student]['Details']}
            for i in ranks[::-1]:
                if student_averages[school][student][i]>=threshold and i != ranks[0]:
                    student_ranks[school][student]['Rank']=i
                    break
                elif i == ranks[0]:
                    student_ranks[school][student]['Rank']=i
                
    return student_ranks, student_averages, student_scores, ranks
''' A dictionary of the form student_ranks = {school 1:{\
                                                                                student 1:{\
                                                                                           rank:___},
                                                                                student 2:{\
                                                                                            rank:___},
                                                                                    .
                                                                                    .
                                                                                    .}
                                                                        school 2:{
                                                                            .
                                                                            .
                                                                            .}}
'''
# rank the schools
def rank_schools(student_ranks,ranks,grade):
    school_ranks={}
    for school in student_ranks:
        school_ranks[school]={}
        for rank in ranks:
            school_ranks[school][rank]=0
        for student in student_ranks[school]:
            school_ranks[school][student_ranks[school][student]['Rank']]+=1
        school_ranks[school]['Number of students']=len(student_ranks[school])
    for school in school_ranks:
        for rank in school_ranks[school]:
            school_ranks[school][rank]=school_ranks[school][rank]/school_ranks[school]['Number of students']*100
        school_ranks[school]['Grade Rank']=school_ranks[school]['G'+str(grade)]+school_ranks[school]['G'+str(grade-1)]
        school_ranks[school]['Number of students']=len(student_ranks[school])
    return school_ranks

#TBD: Read in all Marksheets
#Learner ranks according to Grade Level
threshold = int(input("What's the threshold?"))

maths_data = input_data(input("Maths Sheet Name :"),11,threshold)
language_data = input_data(input("Lang Sheet Name :"),11,threshold)

maths_ranks, maths_averages, maths_scores, maths_ticks =rank_students(maths_data,'Grade Level',threshold)
language_ranks, language_averages, language_scores, language_ticks = rank_students(language_data, 'Grade Level',threshold)        

school_rank_maths=rank_schools(maths_ranks, maths_ticks,8)
school_rank_language=rank_schools(language_ranks,language_ticks,8)

# lets plot some graphs
maths_x = numpy.arange(len(school_rank_maths))
language_x = numpy.arange(len(school_rank_language))

maths_y_prev=[0]*len(maths_x)
for rank in maths_ticks:
    maths_y=[]
    for school in school_rank_maths:
        if rank in school_rank_maths[school]:
            maths_y.append(school_rank_maths[school][rank])
        else:
            maths_y.append(0)
    plt.bar(maths_x,maths_y,bottom=maths_y_prev)
    maths_y_prev = numpy.add(maths_y,maths_y_prev)
plt.show()

language_y_prev=[0]*len(language_x)
for rank in language_ticks:
    language_y=[]
    for school in school_rank_language:
        if rank in school_rank_language[school]:
            language_y.append(school_rank_language[school][rank])
        else:
            language_y.append(0)
    plt.bar(language_x,language_y,bottom=language_y_prev)
    language_y_prev = numpy.add(language_y,language_y_prev)
plt.show()

# Combine language and maths ranks
formatted={}
for school1 in school_rank_maths:
    for school2 in school_rank_language:
        if school1 == school2:
            formatted[school1]={}
            formatted[school1]['Rank']=school_rank_maths[school1]['Grade Rank'] + school_rank_language[school1]['Grade Rank']
            for rank in school_rank_maths[school1]:
                formatted[school1]['M'+str(rank)]=school_rank_maths[school1][rank]
            for rank in school_rank_language[school1]:
                formatted[school1]['L'+str(rank)]=school_rank_language[school1][rank]

        elif school1 not in school_rank_language:
            formatted[school1]={}
            formatted[school1]['Rank']=school_rank_maths[school1]['Grade Rank']
            for rank in school_rank_maths[school1]:
                formatted[school1]['M'+str(rank)]=school_rank_maths[school1][rank]
            for rank in school_rank_language[school2]:
                formatted[school1]['L'+str(rank)]=None

        elif school2 not in school_rank_maths:
            formatted[school2]={}
            formatted[school2]['Rank']=school_rank_language[school2]['Grade Rank']
            for rank in school_rank_maths[school1]:
                formatted[school2]['M'+str(rank)]=0
            for rank in school_rank_language[school2]:
                formatted[school2]['L'+str(rank)]=school_rank_language[school2][rank]



    


# Bubble student averages to the school level
school_averages={}
for school1 in maths_averages:
    for school2 in language_averages:
        if school1==school2:
            school_averages[school1]={}
            for student in maths_averages[school1]:
                for grade in maths_averages[school1][student]:
                    if grade!='Details':
                        if 'M'+grade not in school_averages[school1]:
                            school_averages[school1]['M'+grade]=maths_averages[school1][student][grade]/school_rank_maths[school1]['Number of students']
                        else:
                            school_averages[school1]['M'+grade]+=maths_averages[school1][student][grade]/school_rank_maths[school1]['Number of students']
            for student in language_averages[school2]:    
                for grade in language_averages[school2][student]:
                    if grade !='Details':
                        if 'L'+grade not in school_averages[school2]:
                            school_averages[school2]['L'+grade]=language_averages[school1][student][grade]/school_rank_language[school1]['Number of students']
                        else:
                            school_averages[school2]['L'+grade]+=language_averages[school1][student][grade]/school_rank_language[school1]['Number of students']
            
        elif school1 not in language_averages:
            school_averages[school1]={}
            for student in maths_averages[school1]:
                for grade in maths_averages[school1][student]:
                    if grade != 'Details':
                        if 'M'+grade not in school_averages[school1]:
                            school_averages[school1]['M'+grade]=maths_averages[school1][student][grade]/school_rank_maths[school1]['Number of students']
                        else:
                            school_averages[school1]['M'+grade]+=maths_averages[school1][student][grade]/school_rank_maths[school1]['Number of students']
            for learner in language_averages[school2]:
                for grade in language_averages[school2][learner]:
                    if grade!='Details':
                        school_averages[school1]['L'+grade]=0
                    
        elif school2 not in maths_averages:
            school_averages[school2]={}
            for student in language_averages[school2]:
                for grade in language_averages[school2][student]:
                    if grade!='Details':
                        if 'L'+grade not in school_averages[school1]:
                            school_averages[school2]['L'+grade]=maths_averages[school2][student][grade]/school_rank_language[school2]['Number of students']
                        else:
                            school_averages[school2]['L'+grade]+=maths_averages[school2][student][grade]/school_rank_language[school2]['Number of students']
            for learner in maths_averages[school1]:
                for grade in maths_averages[school1][learner]:
                    if grade!='Details':
                        school_averages[school2]['M'+grade]=0        
            

# Learner ranks according to cognitive and content domains
cognitive_maths_ranks, cognitive_maths_averages, cognitive_maths_scores, cognitive_maths_ticks =rank_students(maths_data,'Cognitive Domain',threshold)
cognitive_language_ranks, cognitive_language_averages, cognitive_language_scores, cognitive_language_ticks = rank_students(language_data, 'Cognitive Domain', threshold)    

content_maths_ranks, content_maths_averages, content_maths_scores, content_maths_ticks =rank_students(maths_data,'Content Domain',threshold)
content_language_ranks, content_language_averages, content_language_scores, content_language_ticks = rank_students(language_data, 'Content Domain', threshold)

# Combine maths and language averages for grades and cognitive levels for each learner
student_averages={}
for school1 in maths_averages:
    for school2 in language_averages:
        for student1 in maths_averages[school1]:
            for student2 in language_averages[school2]:
                if student1==student2:
                    
                    student_averages[student1]={'Details':maths_averages[school1][student1]['Details'],\
                                                  'Maths':{'Grade Levels': maths_averages[school1][student1],\
                                                            'Cognitive Domain':cognitive_maths_averages[school1][student1],\
                                                            'Content Domain':content_maths_averages[school1][student1],\
                                                            'Average':{}},\
                                                   'Language':{'Grade Levels': language_averages[school2][student2],\
                                                               'Cognitive Domain': cognitive_language_averages[school2][student2],\
                                                               'Content Domain': content_language_averages[school2][student2],\
                                                               'Average':{}},\
                                                   'Overall Average':{}}
                     

                elif student1 not in student_averages and student1!=student2 and student1 not in language_averages:
                    
                    student_averages[student1]={'Details':maths_averages[school1][student1]['Details'],\
                                                  'Maths':{'Grade Levels': maths_averages[school1][student1],\
                                                            'Cognitive Domain':cognitive_maths_averages[school1][student1],\
                                                            'Content Domain':content_maths_averages[school1][student1],\
                                                            'Average':{}},\
                                                   'Language':{'Grade Levels': {'Details':''},\
                                                               'Cognitive Domain': {'Details':''},\
                                                               'Content Domain': {'Details':''},\
                                                               'Average':{},\
                                                   'Overall Average':{}}}
                    for i in language_ticks:
                        student_averages[student1]['Language']['Grade Levels'][i]=0
                    for i in cognitive_language_ticks:
                          student_averages[student1]['Language']['Cognitive Domain'][i]=0
                    for i in content_language_ticks:
                          student_averages[student1]['Language']['Content Domain'][i]=0
                elif student2 not in student_averages and student1!=student2 and student2 not in maths_averages:
                    
                    student_averages[student2]={'Details':language_averages[school2][student2]['Details'],\
                                                  'Maths':{'Grade Levels':{'Details':''},\
                                                            'Cognitive Domain':{'Details':''},\
                                                            'Content Domain':{'Details':''},\
                                                            'Average':{}},\
                                                   'Language':{'Grade Levels': language_averages[school2][student2],\
                                                               'Cognitive Domain': cognitive_language_averages[school2][student2],\
                                                               'Content Domain': content_language_averages[school2][student2],\
                                                               'Average':{},\
                                                   'Overall Average':{}}}
                    for i in maths_ticks:
                        student_averages[student2]['Maths']['Grade Levels'][i]=0
                    for i in cognitive_maths_ticks:
                          student_averages[student2]['Maths']['Cognitive Domain'][i]=0
                    for i in content_maths_ticks:
                          student_averages[student2]['Maths']['Content Domain'][i]=0
# Calculate Averages
for student in student_averages:
    m_average = 0
    for grade in student_averages[student]['Maths']['Grade Levels']:
        if grade!='Details':
            m_average+=student_averages[student]['Maths']['Grade Levels'][grade]/(len(student_averages[student]['Maths']['Grade Levels'])-1)
    student_averages[student]['Maths']['Average']=m_average
    l_average = 0
    for grade in student_averages[student]['Language']['Grade Levels']:
        if grade!='Details':
            l_average+=student_averages[student]['Language']['Grade Levels'][grade]/(len(student_averages[student]['Language']['Grade Levels'])-1)
    student_averages[student]['Language']['Average']=l_average 
    student_averages[student]['Overall Average'] = (l_average+m_average)/2

#Write data to excel using openpyxl

new_book = openpyxl.Workbook()
sheet1=new_book.active
sheet1.title = 'School Rank'
sheet1.cell(row=1,column=1).value =  'School'
for r,school in enumerate(formatted):
    sheet1.cell(row=r+2,column = 1).value = school
    for c, rank in enumerate(formatted[school]):
        sheet1.cell(row = 1, column = c +2).value = rank
        sheet1.cell(row = r +2, column = c+2).value = formatted[school][rank]
sheet2=new_book.create_sheet('School Averages')
sheet2.cell(row=1,column=1).value= 'School'
for r,school in enumerate(school_averages):
    sheet2.cell(row=r+2,column=1).value=school
    for c,grade in     enumerate(school_averages[school]):
        sheet2.cell(row=1,column=c+2).value=grade
        sheet2.cell(row=r+2,column=c+2).value=school_averages[school][grade]

# Write student average data to excel
sheet3=new_book.create_sheet('Student Averages')
sheet4=new_book.create_sheet('Grade Level Averages')
sheet5=new_book.create_sheet('Cognitive Domain Averages')
sheet6=new_book.create_sheet('Content Domain Averages')

sheet3.cell(row=2,column = 2+len(student_averages[student]['Details'])).value = 'Maths Average'
sheet3.cell(row=2,column = 3+len(student_averages[student]['Details'])).value = 'Language Average'
sheet3.cell(row=2,column = 4+len(student_averages[student]['Details'])).value = 'Overall Average'

for r,student in enumerate(student_averages):
    
    sheet3.cell(row=r+3, column=2+len(student_averages[student]['Details'])).value = student_averages[student]['Maths']['Average']
    sheet3.cell(row=r+3, column=3+len(student_averages[student]['Details'])).value = student_averages[student]['Language']['Average']
    sheet3.cell(row=r+3, column=4+len(student_averages[student]['Details'])).value = student_averages[student]['Overall Average']
    
    
    
    for c,grade in enumerate(student_averages[student]['Maths']['Grade Levels']):
        if grade!='Details':
            sheet4.cell(row=2, column=c+len(student_averages[student]['Details'])+1).value = grade
            sheet4.cell(row=r+3, column=c+len(student_averages[student]['Details'])+1).value = student_averages[student]['Maths']['Grade Levels'][grade]
    for c,grade in enumerate(student_averages[student]['Language']['Grade Levels']):
        if grade!='Details':
            sheet4.cell(row=2, column=c+len(student_averages[student]['Details'])+len(student_averages[student]['Maths']['Grade Levels'])).value = grade
            sheet4.cell(row=r+3, column=c+len(student_averages[student]['Details'])+len(student_averages[student]['Maths']['Grade Levels'])).value = student_averages[student]['Language']['Grade Levels'][grade]
    
    for c,grade in enumerate(student_averages[student]['Maths']['Cognitive Domain']):
        if grade!='Details':
            sheet5.cell(row=2, column=c+len(student_averages[student]['Details'])+1).value = grade
            sheet5.cell(row=r+3, column=c+len(student_averages[student]['Details'])+1).value = student_averages[student]['Maths']['Cognitive Domain'][grade]
    for c,grade in enumerate(student_averages[student]['Language']['Cognitive Domain']):
        if grade!='Details':
            sheet5.cell(row=2, column=c+len(student_averages[student]['Details'])+len(student_averages[student]['Maths']['Cognitive Domain'])).value = grade
            sheet5.cell(row=r+3, column=c+len(student_averages[student]['Details'])+len(student_averages[student]['Maths']['Cognitive Domain'])).value = student_averages[student]['Language']['Cognitive Domain'][grade]
    
    for c,grade in enumerate(student_averages[student]['Maths']['Content Domain']):
        if grade!='Details':
            sheet6.cell(row=2, column=c+len(student_averages[student]['Details'])+1).value = grade
            sheet6.cell(row=r+3, column=c+len(student_averages[student]['Details'])+1).value = student_averages[student]['Maths']['Content Domain'][grade]
    for c,grade in enumerate(student_averages[student]['Language']['Content Domain']):
        if grade!='Details':
            sheet6.cell(row=2, column=c+len(student_averages[student]['Details'])+len(student_averages[student]['Maths']['Content Domain'])).value = grade
            sheet6.cell(row=r+3, column=c+len(student_averages[student]['Details'])+len(student_averages[student]['Maths']['Content Domain'])).value = student_averages[student]['Language']['Content Domain'][grade]        
    for c,detail in enumerate(student_averages[student]['Details']):
        if detail != 'Scores':
            sheet3.cell(row=2, column=c+1).value = detail
            sheet3.cell(row=r+3,column = c+1).value = student_averages[student]['Details'][detail]
            sheet4.cell(row=2, column=c+1).value = detail
            sheet4.cell(row=r+3,column = c+1).value = student_averages[student]['Details'][detail]
            sheet5.cell(row=2, column=c+1).value = detail
            sheet5.cell(row=r+3,column = c+1).value = student_averages[student]['Details'][detail]
            sheet6.cell(row=2, column=c+1).value = detail
            sheet6.cell(row=r+3,column = c+1).value = student_averages[student]['Details'][detail]
        
sheet3.cell(row=2, column=1).value= 'Student'
sheet4.cell(row=2, column=1).value= 'Student'
sheet5.cell(row=2, column=1).value= 'Student'
sheet6.cell(row=2, column=1).value= 'Student'

new_book.save('SchoolRank_' + datetime.now().isoformat(timespec='minutes') + '.xlsx')